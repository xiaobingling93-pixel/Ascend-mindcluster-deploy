#!/usr/bin/env python3
import openpyxl
import csv
import os
import argparse
import sys


def find_sheet_by_substring(wb, substring):
    """Find worksheet containing specified substring in its name"""
    for sheet_name in wb.sheetnames:
        if substring in sheet_name:
            return wb[sheet_name]
    return None


def find_header_cell(sheet, header_text):
    """Find cell position containing specified header text"""
    for row in sheet.iter_rows(min_row=1, max_row=100):  # Assume headers are within first 100 rows
        for cell in row:
            if cell.value and str(cell.value).strip() == header_text:
                return cell.row, cell.column
    return None, None


def extract_column_data(sheet, header_row, header_col, max_rows=20000):
    """Extract data below specified header column"""
    data = []
    # Start reading from the row after the header, stop at empty cell or max_rows
    for row in range(header_row + 1, header_row + max_rows + 1):
        cell_value = sheet.cell(row=row, column=header_col).value
        if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ""):
            # Stop reading when encountering empty cell
            break
        data.append(cell_value)
    return data


def process_excel_to_csv(input_file, output_file, topotree_name, group_per_rack):
    """Process Excel file and generate CSV file"""

    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' does not exist")
        return False

    try:
        # Load workbook
        wb = openpyxl.load_workbook(input_file, data_only=True)

        # Find worksheet containing "超节点规划"
        target_sheet = find_sheet_by_substring(wb, "超节点规划")

        if target_sheet is None:
            print("Error: Worksheet containing '超节点规划' not found")
            return False

        print(f"Target worksheet found: {target_sheet.title}")

        # Find positions of header cells
        hostname_row, hostname_col = find_header_cell(target_sheet, "主机名称")
        frame_row, frame_col = find_header_cell(target_sheet, "机框编号")

        # Check if all headers are found
        missing_headers = []
        if not hostname_row or not hostname_col:
            missing_headers.append("主机名称")
        if not frame_row or not frame_col:
            missing_headers.append("机框编号")

        if missing_headers:
            print(f"Error: The following headers are not found: {', '.join(missing_headers)}")
            return False

        # Extract data from columns
        node_names = extract_column_data(target_sheet, hostname_row, hostname_col, max_rows=20000)
        frame_numbers = extract_column_data(target_sheet, frame_row, frame_col, max_rows=20000)

        # Check if data lengths are consistent
        min_length = min(len(node_names), len(frame_numbers))
        if len(node_names) != len(frame_numbers):
            print(
                f"Warning: Data column lengths inconsistent - Hostnames: {len(node_names)}, Frame Numbers: {len(frame_numbers)}")
            print(f"Using minimum length: {min_length}")

        # Process data: Calculate groupid (frame number divided by group_per_rack, integer quotient)
        group_ids = []
        for i in range(min_length):
            try:
                # Ensure frame_numbers[i] is numeric
                frame_num = float(frame_numbers[i]) if frame_numbers[i] is not None else 0
                group_id = int((frame_num - 1) // group_per_rack)  # Integer division for quotient
                group_ids.append(group_id)
            except (ValueError, TypeError) as e:
                print(
                    f"Warning: Frame number '{frame_numbers[i]}' at row {i + 1} cannot be converted to numeric, using default 0")
                group_ids.append(0)

        # Write to CSV file
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)

            # Write header row
            writer.writerow(['nodeName', 'huawei.com/topotree.groupid', 'huawei.com/topotree'])

            # Write data rows
            for i in range(min_length):
                writer.writerow([
                    node_names[i],
                    group_ids[i],
                    topotree_name
                ])

        print(f"CSV file successfully generated: {output_file}")
        print(f"Processed {min_length} rows of data")
        print(f"Topotree name: {topotree_name}")
        print(f"Group per rack (divisor): {group_per_rack}")

        return True

    except Exception as e:
        print(f"Error during processing: {str(e)}")
        return False


def main():
    """Main function: Handle command line arguments and interactive input"""

    # Set up command line argument parser
    parser = argparse.ArgumentParser(description='Extract data from Excel file and generate CSV file')
    parser.add_argument('--input', '-i', help='Input Excel file path')
    parser.add_argument('--output', '-o', help='Output CSV file path')
    parser.add_argument('--topotree-name', '-t', help='Topotree name')
    parser.add_argument('--group-per-rack', '-g', type=int, default=12,
                        help='Number of frames per group (default: 12). This is used as divisor for calculating groupid from frame number.')

    args = parser.parse_args()

    # Get input file path
    input_file = args.input
    if not input_file:
        input_file = input("Please enter Excel file path: ").strip()
        if not input_file:
            print("Error: Input file path is required")
            sys.exit(1)

    # Get output file path
    output_file = args.output
    if not output_file:
        # Default: use input filename (without extension) plus _output.csv
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{base_name}_output.csv"
        print(f"Output file not specified, using default: {output_file}")

    # Get topotree name
    topotree_name = args.topotree_name
    if not topotree_name:
        topotree_name = input("Please enter topotree name: ").strip()
        if not topotree_name:
            print("Error: Topotree name is required")
            sys.exit(1)

    # Get group-per-rack value
    group_per_rack = args.group_per_rack

    # Validate group_per_rack is positive
    if group_per_rack <= 0:
        print(f"Error: group-per-rack must be a positive integer, got {group_per_rack}")
        sys.exit(1)

    # Process Excel file
    success = process_excel_to_csv(input_file, output_file, topotree_name, group_per_rack)

    if success:
        print("Script execution completed")
    else:
        print("Script execution failed")
        sys.exit(1)


if __name__ == "__main__":
    main()
